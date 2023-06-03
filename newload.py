from bs4 import  BeautifulSoup
import requests
from openpyxl import load_workbook
import openpyxl
from nltk.tokenize import sent_tokenize, word_tokenize
import numpy as np
from nltk.tokenize import RegexpTokenizer  # for removing punctuations
import string
import nltk
nltk.download('punkt')
nltk.download('averaged_perceptron_tagger')
from nltk.corpus import cmudict
import re
import os


# # enter the path for the Input.xlsx file 
# df = pd.read_excel('Input.xlsx', sheet_name='Sheet1')

# # Extract the URL column as a list
# url_id = df['URL_ID'].tolist()
# url_link = df['URL'].tolist()

# # Convert each URL to a string and append it to a new list
# url_strings = []
# for url_l in url_link:
#     url_strings.append(str(url_l))

# url_id_strings = []

# for url_i in url_id:
#     url_id_strings.append(str(url_i))
# i=0

# file_name = url_id_strings[i]
# for item in url_strings:
#     file_name = url_id_strings[i]
#     i+=1
#     print(f'{file_name}: {item}')
#     check_url = item
#     try:
#         response = requests.get(check_url)
#         if response.status_code == 200:
#             content = response.content
#             print(f"{file_name} is a valid url")
#             html_url = item
#             html_text = requests.get(html_url).text
#             soup = BeautifulSoup(html_text,'lxml')
#             article_title = soup.find('h1',class_ = 'entry-title' )

#             article_text = soup.find('div',class_ = 'td-post-content')
#             # please enter the path of the complete path folder where you want to save the scrapped files insert the path before '/{file_name}'
#             with open(f'D:/Om data/Intership/Blackcoffer/Task 1/Blackcoffer/content2/{file_name}.txt', 'w',encoding="utf-8") as f:
#                 article_title = soup.find('h1',class_ = 'entry-title' ).text
#                 f.write(f'{article_title.strip()} \n')  
#                 article_text = soup.find('div',class_ = 'td-post-content').text.replace(' "" ',' '' ')
#                 f.write(f'{article_text.strip()} \n ') 
#                 print(f'finished saving file : {file_name}.txt')
#                 print("-------------------------")
#         else: 
#             print(f"{file_name} is not a valid url")
#             print("-------------------------")


#     except requests.exceptions.RequestException as e:
#         print("Error fetching URL:", check_url)
#         print(e)


''' Data Analysis '''   
# Positive_Score = []
# Negative_Score = []
# Polarity_Score = []
# Subjectivity_Score = []
# Average_Sentence_Length = []
# Percentage_of_Complex__words = []
# Fog_Index = []
# Average_Number_of_Words_Per_Sentence = []
# Complex_Word_Count = []
# Word_Count = []
# Syllable_Count_Per_Word = []
# Personal_Pronouns = []
Average_Word_Length = set()

# workbook = load_workbook('Output Data Structure.xlsx')
# worksheet = workbook.active

# set folder path for the scrapped data saved directory before running the code
# the folder containing all the txt files
folder_path = "D:/Om data/Intership/Blackcoffer/Task 1/Blackcoffer/content2"
# iterate over files in folder
for filename in os.listdir(folder_path):
    print(filename)
    name_of_file = filename
    # read text file
    path = (folder_path + '/'+ filename)
    print(path)
    with open(path, 'r' ,encoding="utf-8") as file:
        text_file = file.read()
        #tokenize
        words = nltk.word_tokenize(text_file)
        sent = nltk.sent_tokenize(text_file)
        # Remove punctuation from each word
        words = [word for word in words if word not in string.punctuation]
        sent = [word for word in sent if word not in string.punctuation]
        
        ''' Personal Pronouns '''
        def count_personal_pronouns(text):
            # define regex pattern for personal pronouns
            pattern = r"\b(I|we|my|ours|us)\b"

            # find all matches of pattern in text
            matches = re.findall(pattern, text, re.IGNORECASE)

            # count number of matches excluding "us"
            count = 0
            for match in matches:
                if match.lower() != "US":
                    count += 1
            
            return count

        filename = path  # Replace with the filename of your text file

        with open(filename, 'r' ,encoding="utf-8") as input_file:
            text = input_file.read()
        # print(text)
        count = count_personal_pronouns(text)
        print(f'Number of personal pronouns: {count}')

        ''' Positive Score'''
        i = 1

        # reading the postive dictionary 
        filename_positive ='MasterDictionary\positive-words.txt'
        with open(filename_positive, 'r' ,encoding="utf-8") as input_file:
            test_file_positive = input_file.read()

        #check for words in positive dictionary
        score_p = []
        positive_score = 0
        for items in words:
            if items in test_file_positive:
                positive_score +=1
                score_p.append(items)
            else : 
                None
        
        print(f'postive score : {round(positive_score,3)}')

        # letters.append(positive_score)

        ''' Negative Score'''

        #check for words in negactive dictionary
        filename_negative ='MasterDictionary/negative-words.txt'
        with open(filename_negative, 'r' ,) as input_file:
            test_file_negative = input_file.read()

        score_n = []
        negative_score = 0
        for items in words:
            if items in test_file_negative:
                score_n.append(items)
                negative_score -=1
                nscore = negative_score *-1
            else : 
                None
        # print(f'negative score : {round(nscore,3)}')

        ''' Polarity Score'''

        def polarity_score(positive_score,nscore):
            result = ((positive_score-nscore)/ (positive_score+nscore)) + 0.000001
            print(f'the polarity score = {round(result,3)}')
            return result

        # polarity_score(positive_score,nscore)


        ''' Subjectivity Score'''

        def subjectivity__score(positive_score,nscore):
            result = ((positive_score+nscore))/ ((len(words))+0.000001)
            print(f'the subjectivity score = {round(result,3)}')
            return result

        # subjectivity__score(positive_score,nscore)

        ''' Average Sentence Length '''
        # average sentence lenght 
        def avg_sent_len():
            result = (len(words))/(len(sent))
            print(f'avg sent len : {round(result,3)}')
            return result

        # avg_sent_len()

        ''' Percentage of Complex words '''

        # percentage of complex words 

        # Load the CMU pronunciation dictionary
        prondict = cmudict.dict()

        # Define the function to count syllables
        def count_syllables(word):
        #Count the number of syllables in a word using the CMU pronunciation dictionary.
            try:
                # Get the list of possible pronunciations for the word
                pronunciations = prondict[word.lower()]

                # Count the number of syllables in the first pronunciation
                syllable_count = sum([1 for phoneme in pronunciations[0] if phoneme[-1].isdigit()])

                # Return the syllable count
                return syllable_count
            except KeyError:
                return 0
                # Return 0 if the word is not found in the pronunciation dictionary
        # Initialize the count of complex words
        complex_word_count = 0

        # Loop through each word and count the syllables
        for word in words:
            # Check if the word ends in "es" or "ed"
            if word.endswith("es") or word.endswith("ed"):
                continue
            # Count the syllables in the word
            syllable_count = count_syllables(word)
            # If the word has more than 2 syllables, increment the count of complex words
            if syllable_count > 2:
                complex_word_count += 1

        def per_complex(complex_word_count):
            result = complex_word_count/(len(words))
            print(f'percentage of complex words : {round(result,3)}%')
            return result
        # per_complex(complex_word_count)

        ''' Fog Index '''

        def fog_index():
            result = 0.4 * (avg_sent_len() + per_complex(complex_word_count))
            print(f'fog index : {round(result,3)}')
            return result
        # fog_index()

        ''' Average Number of Words Per Sentence '''

        def avg_no_words_per_sent():
            result = len(words)/len(sent)
            print(f'avg_no_words per sent : {round(result,3)}')
            return result
        # avg_no_words_per_sent()

        ''' Complex Word Count '''

        def com_count(syllable_count):
            result = syllable_count
            print(f'complex words count : {round(result,3)}')
            return result
        # com_count(syllable_count)

        ''' Word Count '''
        with open('D:/Om data/Intership/Blackcoffer/Task 1/Blackcoffer/stopwords.txt', 'r') as stopwords_file:
            stopwords = stopwords_file.read().splitlines()
        with open(path, 'r',encoding="utf-8") as input_file:
            text = input_file.read()
        filtered_text = ' '.join([word for word in text.split() if word.lower() not in stopwords])
        clean_words = nltk.word_tokenize(filtered_text)
        def word_count():
            result = len(clean_words)
            # print(f'word count : {round(result,3)}')
            return result
        # word_count()

        ''' Syllable Count Per Word '''

        def count_syllables_in_file(): # file_path was in function input
        # define vowels
            vowels = "aeiouy"

            # define exceptions
            exceptions = ["es", "ed"]

            file_path = path
            # open file
            with open(file_path,"r",encoding="utf-8") as file:
                text = file.read()

            # split text into words
            words = text.split()

            # count syllables per word
            syllables_per_word = []
            for word in words:
                syllables = 0
                prev_char = ""
                for char in word:
                    if char.lower() in vowels and prev_char.lower() not in vowels:
                        syllables += 1
                    prev_char = char
                for exception in exceptions:
                    if word.endswith(exception):
                        syllables -= 1
                        break
                syllables_per_word.append(syllables)

            # calculate average syllables per word
            total_syllables = sum(syllables_per_word)
            num_words = len(words)
            if num_words > 0:
                avg_syllables_per_word = total_syllables / num_words
            else:
                avg_syllables_per_word = 0

            print(f'avg_syllables_per_word : {round(avg_syllables_per_word,3)}')
            return avg_syllables_per_word
        # count_syllables_in_file()        


                

        ''' Average Word Length '''

        def avg_word_len():
            filename = path
            with open(filename,'r',encoding="utf-8") as f:
                w = [len(word) for line in f for word in line.rstrip().split(" ")]
                w_avg = sum(w)/len(w)
                print(f'avg word length : {round(w_avg,3)}')
        avg_word_len()

        ''' save data to excel'''
        
        # Positive_Score.append(positive_score)
        # Negative_Score.append(nscore)
        # Polarity_Score.append(polarity_score(positive_score,nscore))
        # Subjectivity_Score.append(subjectivity__score(positive_score,nscore))
        # Average_Sentence_Length.append(avg_sent_len())
        # Percentage_of_Complex__words.append(per_complex(complex_word_count))
        # Fog_Index.append(fog_index())
        # Average_Number_of_Words_Per_Sentence.append(avg_no_words_per_sent())
        # Complex_Word_Count.append( com_count(syllable_count) )
        # Word_Count.append(word_count())
        # Syllable_Count_Per_Word.append(count_syllables_in_file())
        # Personal_Pronouns.append(count)
        Average_Word_Length.add(avg_word_len())
        print(Average_Word_Length)
        #file path to output data structure file
        # source_filename = "D:/Om data/Intership/Blackcoffer/Task 1/Blackcoffer/Output Data Structure.xlsx"
        # workbook = openpyxl.load_workbook(source_filename)
        # worksheet = workbook.active
        #saving positive score
        # def save_to_excel(func_list,x):
        #     for index, value in enumerate(func_list):
        #         worksheet.cell( row = index +2 , column = x, value = value)
        #     workbook.save('Output Data Structure.xlsx')
        # save_to_excel(Positive_Score,3)
        # save_to_excel(Negative_Score,4)
        # save_to_excel(Polarity_Score,5)
        # save_to_excel(Subjectivity_Score,6)
        # save_to_excel(Average_Sentence_Length,7)
        # save_to_excel(Percentage_of_Complex__words,8)
        # save_to_excel(Fog_Index,9)
        # save_to_excel(Average_Number_of_Words_Per_Sentence,10)
        # save_to_excel(Complex_Word_Count,11)
        # save_to_excel(Word_Count,12)
        # save_to_excel(Syllable_Count_Per_Word,13)
        # save_to_excel(Personal_Pronouns,14)


        # for index, value in enumerate(Average_Word_Length):
        #     worksheet.cell( row = index +2 , column = 15, value = value)
        #     workbook.save('Output Data Structure.xlsx')
        # print("Output value written")
    
        print(f'Name of the file : {name_of_file}')
        print("----------------------------------------------------------------")



# workbook = Workbook()
# worksheet = workbook.active

# worksheet.append([1,2,3,4,5])
# workbook.save('D:/Om data/Intership/Blackcoffer/Task 1/Blackcoffer/om1.xlsx')

# worksheet.append({'G':1,
#                    'h':1,
#                     'I':1, 
#                     'G':1, 
#                     'J':1, 
#                     'K':1, 
#                     })
# workbook.save('D:/Om data/Intership/Blackcoffer/Task 1/Blackcoffer/om1.xlsx')

# letters = ['a','b','c']
# for index, value in enumerate(letters):
#     worksheet.cell( row = index +2 , column = 7, value = value)
# workbook.save('D:/Om data/Intership/Blackcoffer/Task 1/Blackcoffer/om1.xlsx')

