from bs4 import  BeautifulSoup
import requests
from openpyxl import load_workbook
import pandas as pd


# data_file = 'Task 1\Blackcoffer\Input.xlsx'
# # Load the entire workbook.
# wb = load_workbook(data_file)
# # List all the sheets in the file.
# print("Found the following worksheets:")
# for sheetname in wb.sheetnames:
#     print(f'sheet name : {sheetname}')   
# # Load the entire workbook.
# wb = load_workbook(data_file)
# # Load one worksheet.
# ws = wb['Sheet1']
# all_rows = list(ws.rows)
# for cell in all_rows[0]:
#     None

# urls = []

# # Pull information from specific cells.
# for row in all_rows[0:]:
#     url_id = row[0].value
#     url = row[1].value
    
#     print(f"Url id : {url_id}")
#     urls.append(url)
#     print(f" Url link :{url}")
#     # print('')\
#     print('------------------')

# # print(urls)

# Read the Excel sheet into a Pandas DataFrame
df = pd.read_excel('Task 1\Blackcoffer\Input.xlsx', sheet_name='Sheet1')

# Extract the URL column as a list

url_list = df['URL'].tolist()
url_id = df['URL_ID'].tolist()

# Convert each URL to a string and append it to a new list
url_strings = []
for url in url_list:
    url_strings.append(str(url))
# print(url_strings)


# print(url__id_list)
# print(len(url__id_list))



url__id_list = []
for url_id in url__id_list:
    url__id_list.append(str(url_id))
   


for id_name in url__id_list:
        file_name = id_name 
        for item in url_strings:
            print(item)
            check_url = item
            try:
                response = requests.get(check_url)
                if response.status_code == 200:
                    content = response.content
                    print(f"{file_name} is a valid url")
                    html_url = item
                    print(html_url)
                    html_text = requests.get(html_url).text
                    # print(html_text)
                    soup = BeautifulSoup(html_text,'lxml')
                    # print(soup.prettify)
                    # html_url = requests.get(urls).text
                    # print(html_text)
                    soup = BeautifulSoup(html_text,'lxml')
                    article_title = soup.find('h1',class_ = 'entry-title' )
                    # print(article_title)  
                    article_text = soup.find('div',class_ = 'td-post-content')
                    # print(article_text)
                    with open(f'Task 1/Blackcoffer/content2/{file_name}.txt', 'w') as f:
                        article_title = soup.find('h1',class_ = 'entry-title' ).text
                        f.write(f'{article_title.strip()} \n' )  
                        article_text = soup.find('div',class_ = 'td-post-content').text.replace(' "" ',' '' ')
                        f.write(f'{article_text.strip()} \n ') 
                        print("finished saving file")
            except requests.exceptions.RequestException as e:
                print("Error fetching URL:", check_url)
                print(e)
                










